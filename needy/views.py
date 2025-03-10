from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from jdatetime import datetime as jdatetime
import pandas as pd
from django.http import HttpResponse
from django.shortcuts import get_list_or_404
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl import Workbook
from pandas.io.formats.printing import PrettyDict

from .models import Needy, Street, NeedyPath
from datetime import datetime
import jdatetime  # کتابخانه‌ی تبدیل تاریخ به شمسی


def convert_persian_to_gregorian(persian_date):
    try:
        persian_date_obj = jdatetime.datetime.strptime(persian_date, "%Y/%m/%d")
        gregorian_date = persian_date_obj.togregorian()
        return gregorian_date.strftime("%Y-%m-%d")  # تبدیل به فرمت میلادی YYYY-MM-DD
    except ValueError:
        return None


def convert_gregorian_to_jalali(date_string):
    # تبدیل رشته ورودی به تاریخ میلادی
    gregorian_date = datetime.strptime(str(date_string), "%Y-%m-%d")

    # تبدیل تاریخ میلادی به شمسی
    jalali_date = jdatetime.date.fromgregorian(
        year=gregorian_date.year,
        month=gregorian_date.month,
        day=gregorian_date.day
    )

    # فرمت‌دهی تاریخ شمسی به صورت YYYY/MM/DD
    formatted_jalali_date = jalali_date.strftime("%Y/%m/%d")

    return formatted_jalali_date


from django.shortcuts import render


def render_with_error(request, input_data, message, streets, needy_paths):
    return render(request, 'needy/add_needy.html', {
        'input_data': input_data,
        'selected_path': request.POST.get('path', 'undefined'),
        'path_choices': Needy.PATH_CHOICES,
        "coverage_choices": Needy.COVERAGE_CHOICES,
        'selected_marital_status': Needy.MARITAL_STATUS_CHOICES,
        'religion_choices': Needy.RELIGION_CHOICES,
        'streets': streets,
        'needy_paths': needy_paths,
        'message': message
    })


def add_needy(request):
    if not request.user.is_authenticated:
        return redirect('login')

    streets = Street.objects.all()
    needy_paths = NeedyPath.objects.all()
    input_data = {}

    if request.method == 'POST':
        user = request.user
        has_introducer = request.POST.get('has_introducer')
        introducer_name = request.POST.get('introducer_name')
        introducer_phone = request.POST.get('introducer_phone')
        full_name = request.POST.get('full_name')
        father_name = request.POST.get('father_name')
        family_members = request.POST.get('family_members')
        birth_date = request.POST.get('birth_date')
        marital_status = request.POST.get('marital_status')
        religion = request.POST.get('religion')
        job = request.POST.get('job')
        is_covered = request.POST.get('is_covered')
        national_code = request.POST.get('national_code')
        phone_number = request.POST.get('phone_number')
        street = request.POST.get('street')
        path = request.POST.get('path')
        address = request.POST.get('address')
        house_number = request.POST.get('house_number')
        description = request.POST.get('description')

        input_data = {
            'has_introducer': has_introducer,
            'introducer_name': introducer_name,
            'introducer_phone': introducer_phone,
            'full_name': full_name,
            'father_name': father_name,
            'family_members': family_members,
            'birth_date': birth_date,
            'marital_status': marital_status,
            'religion': religion,
            'job': job,
            'is_covered': is_covered,
            'national_code': national_code,
            'phone_number': phone_number,
            'street': street,
            'path': path,
            'address': address,
            'house_number': house_number,
            'description': description,
        }

        birth_date = convert_persian_to_gregorian(birth_date)
        if not street:
            return render_with_error(request, input_data,
                                     'خیابان انتخاب شده یافت نشد. لطفاً یک خیابان معتبر انتخاب کنید.', streets,
                                     needy_paths)

        if not path:
            return render_with_error(request, input_data,
                                     'مسیر انتخابی برای خیابان موجود نیست. لطفاً یک مسیر معتبر برای خیابان انتخاب کنید.',
                                     streets, needy_paths)

        try:
            street_instance = Street.objects.get(id=street)
            path_instance = NeedyPath.objects.get(street=street_instance, id=path)
        except Street.DoesNotExist:
            return render_with_error(request, input_data,
                                     'خیابان انتخاب شده یافت نشد. لطفاً یک خیابان معتبر انتخاب کنید.', streets,
                                     needy_paths)
        except NeedyPath.DoesNotExist:
            return render_with_error(request, input_data,
                                     'مسیر انتخابی برای خیابان موجود نیست. لطفاً یک مسیر معتبر برای خیابان انتخاب کنید.',
                                     streets, needy_paths)
        except:
            return render_with_error(request, input_data, 'خطا در اطلاعات ورودی.', streets, needy_paths)

        if has_introducer is None:
            introducer_name = user.get_full_name()
            introducer_phone = user.username

        try:
            Needy.objects.create(
                created_by=user,
                introducer_name=introducer_name,
                introducer_phone=introducer_phone,
                full_name=full_name,
                father_name=father_name,
                family_members=family_members,
                birth_date=birth_date,
                marital_status=marital_status,
                religion=religion,
                job=job,
                is_covered=is_covered,
                national_code=national_code,
                phone_number=phone_number,
                street=street_instance,
                address=address,
                house_number=house_number,
                description=description,
                path=path_instance
            )
            return redirect('success_view')
        except Exception as e:
            print(str(e))

    return render(request, 'needy/add_needy.html', {
        'selected_path': request.POST.get('path', 'undefined'),
        'input_data': input_data,  # ارسال داده‌ها به فرم برای پر کردن فیلدها
        'path_choices': Needy.PATH_CHOICES,
        "coverage_choices": Needy.COVERAGE_CHOICES,
        'selected_marital_status': Needy.MARITAL_STATUS_CHOICES,
        'religion_choices': Needy.RELIGION_CHOICES,
        'streets': streets,
        'needy_paths': needy_paths
    })


def success_view(request):
    if not request.user.is_authenticated:
        return redirect('login')
    last_needy = Needy.objects.filter(created_by=request.user).last()

    if last_needy is None:
        return redirect('register_needy')

    return render(request, 'needy/success.html', {'needy': last_needy})


def needy_list(request):
    if not request.user.is_authenticated:
        return redirect('login')
    needy = Needy.objects.all()
    context = {'needy': needy}
    return render(request, 'needy/needy_list.html', context)


from django.shortcuts import render, get_object_or_404, redirect
from .models import Needy


def convert_to_persian_numbers(date_str):
    # اعداد فارسی
    persian_digits = {'0': '۰', '1': '۱', '2': '۲', '3': '۳', '4': '۴', '5': '۵', '6': '۶', '7': '۷', '8': '۸',
                      '9': '۹'}

    # تبدیل اعداد به اعداد فارسی
    persian_date = ''.join([persian_digits.get(char, char) for char in date_str])

    return persian_date


def edit_needy(request, needy_id):
    if not request.user.is_authenticated:
        return redirect('login')

    needy = get_object_or_404(Needy, id=needy_id)
    streets = Street.objects.all()
    needy_paths = NeedyPath.objects.all()
    input_data = {}

    import jdatetime

    jalali_date = jdatetime.date.fromgregorian(date=needy.birth_date).strftime("%Y/%m/%d") if needy.birth_date else ""
    if jalali_date :
        jalali_date=converted_date = convert_to_persian_numbers(jalali_date)
    if request.method == 'POST':
        has_introducer = request.POST.get('has_introducer')
        introducer_name = request.POST.get('introducer_name')
        introducer_phone = request.POST.get('introducer_phone')
        full_name = request.POST.get('full_name')
        father_name = request.POST.get('father_name')
        family_members = request.POST.get('family_members')
        birth_date = request.POST.get('birth_date')
        marital_status = request.POST.get('marital_status')
        religion = request.POST.get('religion')
        job = request.POST.get('job')
        is_covered = request.POST.get('is_covered') == 'on'
        national_code = request.POST.get('national_code')
        phone_number = request.POST.get('phone_number')
        street = request.POST.get('street', '').strip()
        path = request.POST.get('path', '').strip()
        address = request.POST.get('address')
        house_number = request.POST.get('house_number')
        description = request.POST.get('description')

        input_data = {
            'has_introducer': has_introducer,
            'introducer_name': introducer_name,
            'introducer_phone': introducer_phone,
            'full_name': full_name,
            'father_name': father_name,
            'family_members': family_members,
            'birth_date': birth_date,
            'marital_status': marital_status,
            'religion': religion,
            'job': job,
            'is_covered': is_covered,
            'national_code': national_code,
            'phone_number': phone_number,
            'street': street,
            'path': path,
            'address': address,
            'house_number': house_number,
            'description': description,
        }

        # تبدیل تاریخ شمسی به میلادی
        birth_date = convert_persian_to_gregorian(birth_date) if birth_date else None

        if not street:
            return render(request, 'needy/edit_needy.html', {
                'error': 'خیابان انتخاب شده یافت نشد. لطفاً یک خیابان معتبر انتخاب کنید.',
                'input_data': input_data,
                'streets': streets,
                'needy_paths': needy_paths,
                'jalali_birth_date': jalali_date,
            })

        if not path:
            return render(request, 'needy/edit_needy.html', {
                'error': 'مسیر انتخابی برای خیابان موجود نیست. لطفاً یک مسیر معتبر برای خیابان انتخاب کنید.',
                'input_data': input_data,
                'streets': streets,
                'needy_paths': needy_paths,
                'jalali_birth_date': jalali_date,
            })

        try:
            street_instance = Street.objects.get(id=street)
            path_instance = NeedyPath.objects.get(street=street_instance, id=path)
        except (Street.DoesNotExist, NeedyPath.DoesNotExist):
            return render(request, 'needy/edit_needy.html', {
                'error': 'خیابان یا مسیر انتخاب شده نامعتبر است.',
                'input_data': input_data,
                'streets': streets,
                'needy_paths': needy_paths,
                'jalali_birth_date': jalali_date,
            })

        needy.full_name = full_name
        needy.father_name = father_name
        needy.family_members = family_members
        needy.birth_date = birth_date
        needy.marital_status = marital_status
        needy.religion = religion
        needy.job = job
        needy.is_covered = is_covered
        needy.national_code = national_code
        needy.phone_number = phone_number
        needy.street = street_instance
        needy.path = path_instance
        needy.address = address
        needy.house_number = house_number
        needy.description = description

        needy.save()
        return redirect('needy_list')
    print("+++++++++++="+f"{jalali_date}")
    return render(request, 'needy/edit_needy.html', {
        'needy': needy,
        'input_data': {
            'full_name': needy.full_name,
            'father_name': needy.father_name,
            'family_members': needy.family_members,
            'birth_date': needy.birth_date,
            'marital_status': needy.marital_status,
            'religion': needy.religion,
            'job': needy.job,
            'is_covered': needy.is_covered,
            'national_code': needy.national_code,
            'phone_number': needy.phone_number,
            'street': needy.street.id if needy.street else '',
            'path': needy.path.id if needy.path else '',
            'address': needy.address,
            'house_number': needy.house_number,
            'description': needy.description,
        },
        'streets': streets,
        'needy_paths': needy_paths,
        'jalali_birth_date': jalali_date,
    })

def delete_needy(request, id):
    needy = get_object_or_404(Needy, id=id)
    needy.delete()
    messages.success(request, 'نیازمند با موفقیت حذف شد.')
    return redirect('needy_list')


def export_needy_to_excel(request):
    if not request.user.is_authenticated:
        return redirect('login')
    if not request.user.is_superuser:
        return redirect('add_needy')

    # دریافت تمامی نیازمندان از دیتابیس
    needy_list = get_list_or_404(Needy)

    # ایجاد یک DataFrame از لیست نیازمندان
    data = {
        'کاربر ایجاد کننده': [needy.created_by.get_full_name() if needy.created_by else 'نامشخص' for needy in
                              needy_list],
        'نام و نام خانوادگی معرف': [needy.introducer_name for needy in needy_list],
        'شماره تلفن معرف': [needy.introducer_phone for needy in needy_list],
        'نام و نام خانوادگی': [needy.full_name for needy in needy_list],
        'نام پدر': [needy.father_name for needy in needy_list],
        'تعداد اعضای تحت سرپرستی': [needy.family_members for needy in needy_list],
        'تاریخ تولد': [convert_gregorian_to_jalali(needy.birth_date) if needy.birth_date else '' for needy in
                       needy_list],  # تبدیل تاریخ به شمسی
        'وضعیت تاهل': [needy.marital_status for needy in needy_list],
        'مذهب': [needy.religion for needy in needy_list],
        'شغل': [needy.job for needy in needy_list],
        'تحت پوشش': [needy.is_covered for needy in needy_list],
        'کد ملی': [needy.national_code for needy in needy_list],
        'شماره تماس': [needy.phone_number for needy in needy_list],
        'مسیر': [needy.path.name if needy.path else '' for needy in needy_list],
        'خیابان': [needy.street.name if needy.street else '' for needy in needy_list],
        'آدرس': [needy.address for needy in needy_list],
        'توضیحات': [needy.description for needy in needy_list],
        'موقعیت جغرافیایی': [needy.location for needy in needy_list],
        'لینک موقعیت مکانی': [needy.location_link for needy in needy_list],
    }

    df = pd.DataFrame(data)

    # ایجاد یک Workbook جدید با openpyxl
    wb = Workbook()
    ws = wb.active

    # تنظیم جهت صفحه به راست‌به‌چپ
    ws.sheet_view.rightToLeft = True

    # اضافه کردن تیترهای ستون‌ها
    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = Font(bold=True)  # تیترهای ستون‌ها را پررنگ می‌کنیم
        cell.alignment = Alignment(horizontal='right')  # تراز راست برای تیترها

    # اضافه کردن داده‌های DataFrame به Worksheet
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):  # شروع از ردیف ۲ به بعد
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # تنظیم تراز متن به راست برای تمام سلول‌ها
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='right')

    # اعمال Border برای تمام سلول‌ها
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    # ایجاد یک پاسخ HTTP با نوع محتوای اکسل
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="needy_list.xlsx"'

    # ذخیره Workbook در فایل اکسل و ارسال آن به عنوان پاسخ
    wb.save(response)

    return response
